"""
merge-lee-export.py

Merges notes from Lee's Export.xlsx "Next Steps - History" into
All-Client-Data-Updated.xlsx "History Notes2", deduplicating by
(normalized name + date + note text).

Lee's Export has NO header row: col A=name, col B=notes, col C=date.

Run from the project root:
    python3 scripts/merge-lee-export.py

Requires: openpyxl
    pip install openpyxl
"""

import os
import base64
import datetime
import re
import openpyxl

BASE_DIR    = os.path.join(os.path.dirname(__file__), "..", "Client Data")
SOURCE_FILE = os.path.join(BASE_DIR, "All-Client-Data-Updated.xlsx")
LEE_FILE    = os.path.join(BASE_DIR, "Lee's Export.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "All-Client-Data-Updated.xlsx")  # overwrite in place


def normalize_name(raw):
    if not raw:
        return ""
    return " ".join(str(raw).strip().split()).title()


def normalize_date(raw):
    if not raw:
        return ""
    if isinstance(raw, (datetime.datetime, datetime.date)):
        return str(raw)[:10]
    s = str(raw).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return s


def normalize_notes(raw):
    if not raw:
        return ""
    return " ".join(str(raw).strip().split())


def to_iso(raw):
    if not raw:
        return None
    if isinstance(raw, datetime.datetime):
        return raw.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    if isinstance(raw, datetime.date):
        return datetime.datetime(raw.year, raw.month, raw.day).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    s = str(raw).strip()
    # Already ISO-ish
    if re.match(r"^\d{4}-\d{2}-\d{2}T", s):
        return s
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10] + "T00:00:00.000Z"
    return s


def generate_row_id():
    return base64.urlsafe_b64encode(os.urandom(16)).decode().rstrip("=")


def main():
    print(f"Loading {SOURCE_FILE} ...")
    wb = openpyxl.load_workbook(SOURCE_FILE)
    clients_ws = wb["Clients"]
    history_ws = wb["History Notes2"]

    # Build name → client ID lookup from Clients sheet
    client_id_lookup = {}
    for row in clients_ws.iter_rows(min_row=2, values_only=True):
        name = normalize_name(row[0])       # col A
        cid  = str(row[1]) if row[1] else None  # col B (encoded ID)
        if name and cid:
            client_id_lookup[name] = cid
    print(f"  {len(client_id_lookup)} clients in lookup")

    # Build dedup set from existing History Notes2 rows
    # Key: (normalized_name, normalized_date, normalized_notes_first_80_chars)
    existing = set()
    for row in history_ws.iter_rows(min_row=2, values_only=True):
        name  = normalize_name(row[10])    # col K: Filter (client name)
        date  = normalize_date(row[11])    # col L: Modified / Date
        notes = normalize_notes(row[2])    # col C: Value
        if name or notes:
            existing.add((name, date, notes[:80]))
    print(f"  {len(existing)} existing notes indexed for dedup")

    # Read Lee's Export - NO header row
    print(f"\nLoading {LEE_FILE} ...")
    wb_lee = openpyxl.load_workbook(LEE_FILE, data_only=True)
    ws_lee = wb_lee["Next Steps - History"]

    appended = 0
    skipped_dup = 0
    skipped_empty = 0
    unmatched_clients = set()

    for row in ws_lee.iter_rows(min_row=1, values_only=True):
        raw_name  = row[0]
        raw_notes = row[1]
        raw_date  = row[2]

        name  = normalize_name(raw_name)
        notes = str(raw_notes).strip() if raw_notes else ""
        date  = normalize_date(raw_date)

        if not name or (not notes and not date):
            skipped_empty += 1
            continue

        # Dedup check
        key = (name, date, normalize_notes(notes)[:80])
        if key in existing:
            skipped_dup += 1
            continue

        # Look up client ID
        client_id = client_id_lookup.get(name)
        if not client_id:
            unmatched_clients.add(name)
            # Still add the note — use name in col K, leave col B blank

        iso_date = to_iso(raw_date)

        new_row = [
            generate_row_id(),  # A: Row ID
            client_id,          # B: Client / ID (may be None if unmatched)
            notes,              # C: Value
            None,               # D
            None,               # E
            None,               # F
            None,               # G
            None,               # H
            None,               # I
            None,               # J
            name,               # K: Filter (client name)
            iso_date,           # L: Modified / Date
            None,               # M: Modified / By (Lee's export has no coach column)
        ]
        history_ws.append(new_row)
        existing.add(key)  # prevent re-adding if Lee's has duplicates internally
        appended += 1

    print(f"\nSaving to {OUTPUT_FILE} ...")
    wb.save(OUTPUT_FILE)

    print("\n--- Summary ---")
    print(f"Lee's Export rows read : {ws_lee.max_row}")
    print(f"Notes appended         : {appended}")
    print(f"Skipped (duplicates)   : {skipped_dup}")
    print(f"Skipped (empty)        : {skipped_empty}")
    print(f"Unmatched client names : {len(unmatched_clients)} (notes still added, Client/ID left blank)")
    if unmatched_clients:
        print("\nUnmatched names:")
        for n in sorted(unmatched_clients):
            print(f"  - {n}")
    print("\nDone. Run generate-all-client-csv.py to regenerate CSVs.")


if __name__ == "__main__":
    main()
