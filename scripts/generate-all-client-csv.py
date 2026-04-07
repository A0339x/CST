"""
generate-all-client-csv.py

Reads All-Client-Data-Updated.xlsx and outputs two CSVs to sheets-export/:
  - clients.csv     : one row per active client (Name, Email, Offer, Coach, Next Step, Last Contact Date)
  - notes-log.csv   : all history notes (Client Name, Date, Notes, Coach)

These CSVs can be imported directly into Google Sheets as an interim CST.
See scripts/SHEETS-SETUP.md for import instructions.

Run from the project root:
    python3 scripts/generate-all-client-csv.py

Requires: openpyxl
    pip install openpyxl
"""

import csv
import re
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

ROOT      = Path(__file__).parent.parent
XLSX_PATH = ROOT / "Client Data" / "All-Client-Data-Updated.xlsx"
OUT_DIR   = ROOT / "sheets-export"

# ---------------------------------------------------------------------------
# Helpers (shared conventions with generate-sheets-csv.py)
# ---------------------------------------------------------------------------

def normalize_name(raw):
    if raw is None:
        return ""
    return " ".join(str(raw).strip().split()).title()


def to_iso(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return s


def cell_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Clients sheet column indices (0-based)
# All-Client-Data-Updated.xlsx "Clients" sheet
# ---------------------------------------------------------------------------
# Col A  (0)  : Client / Name
# Col E  (4)  : Client / Email
# Col Y  (24) : Next Steps / Value
# Col AG (32) : CSM / Date of Last Contact
# Col BN (65) : Program / Status Value  ("front-end" = active)

COL_NAME              = 0
COL_EMAIL             = 4
COL_NEXT_STEP         = 24
COL_NEXT_STEP_DATE    = 25   # Next Steps / (Update Time)
COL_NEXT_STEP_COACH   = 26   # Next Steps / (Update By)
COL_LAST_CONTACT      = 32
COL_STATUS            = 65   # "front-end" = active client

# ---------------------------------------------------------------------------
# History Notes2 column indices (0-based)
# ---------------------------------------------------------------------------
# Col C  (2)  : Value (note text)
# Col K  (10) : Filter (client name)
# Col L  (11) : Modified / Date
# Col M  (12) : Modified / By (coach)

COL_NOTE_TEXT   = 2
COL_NOTE_NAME   = 10
COL_NOTE_DATE   = 11
COL_NOTE_COACH  = 12


def parse_clients(sheet) -> dict:
    """
    Returns: { normalized_name -> { email, next_step, last_contact } }
    Only includes clients with Program/Status = 'front-end' (active).
    """
    clients = {}
    skipped = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = normalize_name(row[COL_NAME])
        if not name:
            continue

        # Only include active (front-end) clients
        status = cell_str(row[COL_STATUS]) if len(row) > COL_STATUS else None
        if status and status.lower() not in ("front-end", "front end", "active"):
            skipped += 1
            continue

        clients[name] = {
            "email":           cell_str(row[COL_EMAIL])             if len(row) > COL_EMAIL           else None,
            "next_step":       cell_str(row[COL_NEXT_STEP])         if len(row) > COL_NEXT_STEP       else None,
            "next_step_date":  to_iso(row[COL_NEXT_STEP_DATE])      if len(row) > COL_NEXT_STEP_DATE  else None,
            "next_step_coach": cell_str(row[COL_NEXT_STEP_COACH])   if len(row) > COL_NEXT_STEP_COACH else None,
            "last_contact":    to_iso(row[COL_LAST_CONTACT])        if len(row) > COL_LAST_CONTACT    else None,
        }

    print(f"  -> {len(clients)} active clients parsed ({skipped} non-active skipped)")
    return clients


def parse_history(sheet) -> dict:
    """
    Returns: { normalized_name -> [ {date, notes, coach}, ... ] } sorted oldest→newest.
    Uses col K (Filter) as the client name — matches the pattern set in merge-notes.py.
    """
    history: dict = defaultdict(list)
    skipped = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name  = normalize_name(row[COL_NOTE_NAME]) if len(row) > COL_NOTE_NAME else ""
        notes = cell_str(row[COL_NOTE_TEXT])        if len(row) > COL_NOTE_TEXT else None
        dt    = to_iso(row[COL_NOTE_DATE])           if len(row) > COL_NOTE_DATE else None
        coach = cell_str(row[COL_NOTE_COACH])        if len(row) > COL_NOTE_COACH else None

        if not name:
            skipped += 1
            continue
        if not notes and not dt:
            continue

        history[name].append({"date": dt, "notes": notes, "coach": coach})

    for entries in history.values():
        entries.sort(key=lambda e: e["date"] or "")

    total = sum(len(v) for v in history.values())
    print(f"  -> {total} notes across {len(history)} clients ({skipped} rows skipped — no client name)")
    return dict(history)


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"File not found: {XLSX_PATH}\nRun merge-notes.py first to generate All-Client-Data-Updated.xlsx")

    print(f"Loading: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    print("Parsing Clients sheet...")
    clients = parse_clients(wb["Clients"])

    print("\nParsing History Notes2 sheet...")
    history = parse_history(wb["History Notes2"])

    # -----------------------------------------------------------------------
    # Build clients.csv rows
    # -----------------------------------------------------------------------
    all_names = sorted(set(clients.keys()) | set(history.keys()))
    clients_rows = []

    for name in all_names:
        cd      = clients.get(name, {})
        entries = history.get(name, [])

        # Next Step: from Clients sheet col Y, fallback to most recent note
        next_step = cd.get("next_step")
        if not next_step:
            for e in reversed(entries):
                if e.get("notes"):
                    next_step = e["notes"]
                    break

        # Last Contact Date: from Clients sheet col AG, fallback to most recent dated note
        last_contact = cd.get("last_contact")
        if not last_contact:
            for e in reversed(entries):
                if e.get("date"):
                    last_contact = e["date"]
                    break

        # Coach: from most recent note that has a coach name
        coach = None
        for e in reversed(entries):
            if e.get("coach"):
                coach = e["coach"]
                break

        clients_rows.append({
            "Name":              name,
            "Email":             cd.get("email") or "",
            "Offer":             "",   # not stored in source data
            "Coach":             coach or "",
            "Next Step":         next_step or "",
            "Last Contact Date": last_contact or "",
        })

    # -----------------------------------------------------------------------
    # Build notes-log.csv rows
    # Include the current Next Step from the Clients sheet as the most recent
    # note entry if it isn't already present in History Notes2.
    # -----------------------------------------------------------------------
    next_step_added = 0
    notes_rows = []
    for name in all_names:
        cd      = clients.get(name, {})
        entries = list(history.get(name, []))

        current_note  = cd.get("next_step")
        if current_note:
            existing_texts = {(e.get("notes") or "").strip() for e in entries}
            if current_note.strip() not in existing_texts:
                entries.append({
                    "date":  cd.get("next_step_date") or "",
                    "notes": current_note,
                    "coach": cd.get("next_step_coach") or "",
                })
                next_step_added += 1

        # Re-sort oldest → newest after potential insertion
        entries.sort(key=lambda e: e.get("date") or "")

        for e in entries:
            notes_rows.append({
                "Client Name": name,
                "Date":        e.get("date")  or "",
                "Notes":       e.get("notes") or "",
                "Coach":       e.get("coach") or "",
            })

    # -----------------------------------------------------------------------
    # Write CSVs
    # -----------------------------------------------------------------------
    OUT_DIR.mkdir(exist_ok=True)

    clients_path = OUT_DIR / "clients.csv"
    with open(clients_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Name", "Email", "Offer", "Coach", "Next Step", "Last Contact Date"])
        writer.writeheader()
        writer.writerows(clients_rows)
    print(f"\nWrote {len(clients_rows)} clients  -> {clients_path}")

    notes_path = OUT_DIR / "notes-log.csv"
    with open(notes_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Client Name", "Date", "Notes", "Coach"])
        writer.writeheader()
        writer.writerows(notes_rows)
    print(f"Wrote {len(notes_rows)} notes    -> {notes_path}  ({next_step_added} current next-steps added from Clients sheet)")

    print(f"\nDone. Import both CSVs into Google Sheets per scripts/SHEETS-SETUP.md")


if __name__ == "__main__":
    main()
