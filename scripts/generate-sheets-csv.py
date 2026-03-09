"""
generate-sheets-csv.py
Reads Client-data-updated.xlsx and outputs two CSVs to sheets-export/:
  - clients.csv     : one row per client (Name, Email, Offer, Coach, Next Step, Last Contact Date)
  - notes-log.csv   : one row per note entry (Client Name, Date, Notes, Coach)

Run from the project root:
    python scripts/generate-sheets-csv.py

Requires: openpyxl
    pip install openpyxl
"""

import csv
import re
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).parent.parent
XLSX_PATH = ROOT / "Client-data-updated.xlsx"
OUT_DIR = ROOT / "sheets-export"

# ---------------------------------------------------------------------------
# Helpers (reused from process-client-data.py)
# ---------------------------------------------------------------------------

def normalize_name(raw) -> str:
    """Strip whitespace and title-case."""
    if raw is None:
        return ""
    return str(raw).strip().title()


def to_iso(value) -> str | None:
    """Convert a cell value to an ISO 8601 string, or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]  # trim to date portion
    return s


def cell_str(cell_value) -> str | None:
    """Return stripped string or None."""
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    return s if s else None


def get_headers(sheet) -> dict[str, int]:
    """Read the first row and return {header_lower: col_index (0-based)}."""
    headers = {}
    for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1))):
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = col_idx
    return headers


def _find_col(headers: dict[str, int], candidates: list[str]) -> int | None:
    """Return the column index for the first matching candidate header name."""
    for c in candidates:
        if c.lower() in headers:
            return headers[c.lower()]
    return None


def _sheet_has_header(sheet) -> bool:
    """
    Heuristic: if the first cell of row 1 looks like a short label (not a name/note),
    treat row 1 as a header. Otherwise, data starts at row 1.
    """
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if not first_row or first_row[0] is None:
        return True
    first_val = str(first_row[0]).strip()
    if len(first_val) > 60 or "\n" in first_val or re.match(r"\d{4}-\d{2}", first_val):
        return False
    for keyword in ["name", "client", "email", "note", "coach", "date", "timestamp", "step"]:
        if keyword in first_val.lower():
            return True
    if re.match(r"^[A-Z][a-z]+(?: [A-Z][a-z]+)*$", first_val) and len(first_val) < 40:
        return False
    return True


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def parse_current_data(sheet) -> dict[str, dict]:
    """
    Returns: { normalized_name -> { email, offer, coach, current_next_step } }
    """
    headers = get_headers(sheet)
    print(f"  Current Data headers: {list(headers.keys())}")

    name_col      = _find_col(headers, ["client / name", "name", "client name", "full name", "client"])
    email_col     = _find_col(headers, ["client / email", "email", "email address"])
    offer_col     = _find_col(headers, ["offer", "product", "program"])
    coach_col     = _find_col(headers, ["coach", "coach name", "assigned coach"])
    next_step_col = _find_col(headers, ["next steps / value", "next steps/value", "next steps", "next step", "value", "notes"])
    updated_col   = _find_col(headers, ["updated date", "updated", "modified date", "modified / date", "date"])

    print(f"  Mapped cols -> name:{name_col} email:{email_col} offer:{offer_col} "
          f"coach:{coach_col} next_step:{next_step_col} updated:{updated_col}")

    clients = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_name = row[name_col] if name_col is not None else None
        name = normalize_name(raw_name)
        if not name:
            continue
        clients[name] = {
            "email":             cell_str(row[email_col])     if email_col is not None     else None,
            "offer":             cell_str(row[offer_col])     if offer_col is not None     else None,
            "coach":             cell_str(row[coach_col])     if coach_col is not None     else None,
            "current_next_step": cell_str(row[next_step_col]) if next_step_col is not None else None,
            "updated_date":      to_iso(row[updated_col])     if updated_col is not None   else None,
        }

    print(f"  -> Parsed {len(clients)} clients from Current Data")
    return clients


def parse_call_history(sheet) -> dict[str, list]:
    """
    Returns: { normalized_name -> [ {date, notes, coach}, ... ] }
    Sorted ascending by date within each client.
    """
    has_header = _sheet_has_header(sheet)
    start_row = 2 if has_header else 1

    if has_header:
        headers = get_headers(sheet)
        print(f"  Call History headers: {list(headers.keys())}")
        name_col  = _find_col(headers, ["client / name", "name", "client name", "full name", "client"])
        notes_col = _find_col(headers, ["notes", "note", "next steps / value", "next steps/value", "next steps", "value", "details"])
        date_col  = _find_col(headers, ["timestamp", "date", "call date", "datetime", "modified / date", "modified/date", "updated date"])
        coach_col = _find_col(headers, ["coach", "coach name", "assigned coach", "modified / by", "modified/by", "updated by"])
    else:
        print("  Call History: no header row detected — using positional columns")
        name_col, notes_col, date_col, coach_col = 0, 1, 2, 3

    print(f"  Mapped cols -> name:{name_col} notes:{notes_col} date:{date_col} coach:{coach_col}")

    history: dict[str, list] = defaultdict(list)
    SKIP_NAMES = {"client name", "name", "client"}

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        raw_name = row[name_col] if name_col is not None else None
        name = normalize_name(raw_name)
        if not name or name.lower() in SKIP_NAMES:
            continue
        date_val = to_iso(row[date_col]) if date_col is not None else None
        notes_val = cell_str(row[notes_col]) if notes_col is not None else None
        # Skip rows with no date and no notes
        if not date_val and not notes_val:
            continue
        entry = {
            "date":  date_val,
            "notes": notes_val,
            "coach": cell_str(row[coach_col]) if coach_col is not None else None,
        }
        history[name].append(entry)

    for name in history:
        history[name].sort(key=lambda e: e["date"] or "")

    total = sum(len(v) for v in history.values())
    print(f"  -> Parsed {total} call entries across {len(history)} clients")
    return dict(history)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not XLSX_PATH.exists():
        sys.exit(f"File not found: {XLSX_PATH}\n"
                 f"Make sure Client-data-updated.xlsx is in the project root.")

    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    # -----------------------------------------------------------------------
    # Identify sheets
    # -----------------------------------------------------------------------
    sheet_names = wb.sheetnames
    current_sheet = None
    history_sheet = None

    for sn in sheet_names:
        sl = sn.lower()
        if "current" in sl:
            current_sheet = wb[sn]
        elif ("next step" in sl and "history" in sl) or ("next step" in sl):
            history_sheet = wb[sn]
        elif "history" in sl or "note" in sl or "call" in sl:
            if history_sheet is None:
                history_sheet = wb[sn]

    if current_sheet is None and len(sheet_names) >= 1:
        current_sheet = wb[sheet_names[0]]
    if history_sheet is None and len(sheet_names) >= 2:
        history_sheet = wb[sheet_names[1]]

    print(f"\nMapped sheets:")
    print(f"  Current Data : {current_sheet.title if current_sheet else 'NOT FOUND'}")
    print(f"  Call History : {history_sheet.title if history_sheet else 'NOT FOUND'}")

    # -----------------------------------------------------------------------
    # Parse
    # -----------------------------------------------------------------------
    print("\nParsing Current Data...")
    current_data = parse_current_data(current_sheet) if current_sheet else {}

    print("\nParsing Call History...")
    call_history = parse_call_history(history_sheet) if history_sheet else {}

    # -----------------------------------------------------------------------
    # Build unified client list
    # -----------------------------------------------------------------------
    all_names: set[str] = set(current_data.keys()) | set(call_history.keys())

    # For each client: derive Next Step and Last Contact Date
    clients_rows = []
    for name in sorted(all_names):
        cd = current_data.get(name, {})
        history = call_history.get(name, [])

        # Next Step = last call note that isn't null (prefer current_next_step if set)
        next_step = cd.get("current_next_step")
        if not next_step:
            for entry in reversed(history):
                if entry.get("notes"):
                    next_step = entry["notes"]
                    break

        # Last Contact Date = most recent dated entry
        last_contact = None
        for entry in reversed(history):
            if entry.get("date"):
                last_contact = entry["date"]
                break

        # Coach: from current data, or first available in history
        coach = cd.get("coach")
        if not coach:
            for entry in history:
                if entry.get("coach"):
                    coach = entry["coach"]
                    break

        clients_rows.append({
            "Name":              name,
            "Email":             cd.get("email") or "",
            "Offer":             cd.get("offer") or "",
            "Coach":             coach or "",
            "Next Step":         next_step or "",
            "Last Contact Date": last_contact or "",
        })

    # -----------------------------------------------------------------------
    # Build notes log — history entries + current next step merged together
    # -----------------------------------------------------------------------
    # Collect all entries per client, then sort and deduplicate
    all_notes: dict[str, list] = {}
    for name in sorted(all_names):
        entries = list(call_history.get(name, []))

        # Add current next step as a note entry if it exists and isn't a duplicate
        cd = current_data.get(name, {})
        current_note = cd.get("current_next_step")
        if current_note:
            existing_notes = {e.get("notes") for e in entries}
            if current_note not in existing_notes:
                entries.append({
                    "date":  cd.get("updated_date") or "",
                    "notes": current_note,
                    "coach": cd.get("coach") or "",
                })

        # Sort oldest→newest
        entries.sort(key=lambda e: e.get("date") or "")
        all_notes[name] = entries

    notes_rows = []
    for name in sorted(all_notes.keys()):
        for entry in all_notes[name]:
            notes_rows.append({
                "Client Name": name,
                "Date":        entry.get("date") or "",
                "Notes":       entry.get("notes") or "",
                "Coach":       entry.get("coach") or "",
            })

    # -----------------------------------------------------------------------
    # Write CSVs
    # -----------------------------------------------------------------------
    OUT_DIR.mkdir(exist_ok=True)

    clients_path = OUT_DIR / "clients.csv"
    with open(clients_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Name", "Email", "Offer", "Coach", "Next Step", "Last Contact Date"])
        writer.writeheader()
        writer.writerows(clients_rows)
    print(f"\nWrote {len(clients_rows)} clients -> {clients_path}")

    notes_path = OUT_DIR / "notes-log.csv"
    with open(notes_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Client Name", "Date", "Notes", "Coach"])
        writer.writeheader()
        writer.writerows(notes_rows)
    print(f"Wrote {len(notes_rows)} note entries -> {notes_path}")

    print(f"\nDone. Import both CSVs into Google Sheets -- see scripts/SHEETS-SETUP.md for instructions.")


if __name__ == "__main__":
    main()
