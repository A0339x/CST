"""
merge-notes.py

Merges recent notes from Most-Recent-Updates.xlsx into All-Client-Data.xlsx.
Appends new rows to the "History Notes2" sheet and saves as All-Client-Data-Updated.xlsx.

Usage:
    pip install openpyxl
    python scripts/merge-notes.py
"""

import os
import base64
import datetime
import openpyxl

BASE_DIR = os.path.join(os.path.dirname(__file__), "..", "Client Data")
SOURCE_FILE = os.path.join(BASE_DIR, "All-Client-Data.xlsx")
UPDATES_FILE = os.path.join(BASE_DIR, "Most-Recent-Updates.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "All-Client-Data-Updated.xlsx")

# Only include notes on or after this date
CUTOFF_DATE = datetime.date(2026, 3, 4)


def normalize_name(name):
    if not name:
        return ""
    return " ".join(str(name).strip().split()).title()


def generate_row_id():
    """Generate a 22-char base64url ID matching existing History Notes2 Row ID format."""
    return base64.urlsafe_b64encode(os.urandom(16)).decode().rstrip("=")


def to_iso(dt):
    """Convert a datetime or date object to ISO 8601 string."""
    if isinstance(dt, datetime.datetime):
        return dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    if isinstance(dt, datetime.date):
        return datetime.datetime(dt.year, dt.month, dt.day).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    return str(dt)


def build_client_id_lookup(clients_ws):
    """
    Build a name -> encoded_id lookup from the Clients sheet.
    Col A = Client Name, Col B = Row ID (encoded client ID).
    """
    lookup = {}
    for row in clients_ws.iter_rows(min_row=2, values_only=True):
        name = normalize_name(row[0])  # Col A
        client_id = row[1]             # Col B (encoded ID)
        if name and client_id:
            lookup[name] = str(client_id)
    return lookup


def read_new_notes(notes_ws):
    """
    Read notes from Most-Recent-Updates "Notes Log" sheet.
    Returns list of dicts: {client_name, date, notes, coach}
    """
    entries = []
    for row in notes_ws.iter_rows(min_row=2, values_only=True):
        client_name = row[0]  # Col A
        date_val    = row[1]  # Col B
        notes       = row[2]  # Col C
        coach       = row[3]  # Col D

        # Skip empty rows
        if not client_name and not notes:
            continue

        # Parse date
        if isinstance(date_val, datetime.datetime):
            entry_date = date_val.date()
        elif isinstance(date_val, datetime.date):
            entry_date = date_val
        else:
            entry_date = None

        entries.append({
            "client_name": normalize_name(client_name),
            "date": entry_date,
            "date_obj": date_val,
            "notes": notes,
            "coach": coach,
        })
    return entries


def main():
    print(f"Loading {SOURCE_FILE} ...")
    wb = openpyxl.load_workbook(SOURCE_FILE)

    print(f"Loading {UPDATES_FILE} ...")
    updates_wb = openpyxl.load_workbook(UPDATES_FILE, data_only=True)

    clients_ws = wb["Clients"]
    history_ws = wb["History Notes2"]
    notes_ws   = updates_wb["Notes Log"]

    # Build name → client ID lookup
    client_lookup = build_client_id_lookup(clients_ws)
    print(f"Loaded {len(client_lookup)} clients from Clients sheet.")

    # Read new notes
    new_notes = read_new_notes(notes_ws)
    print(f"Read {len(new_notes)} note rows from Most-Recent-Updates.")

    appended = 0
    skipped_date = 0
    skipped_name = []

    for entry in new_notes:
        # Filter by cutoff date
        if entry["date"] and entry["date"] < CUTOFF_DATE:
            skipped_date += 1
            continue

        client_id = client_lookup.get(entry["client_name"])
        if not client_id:
            skipped_name.append(entry["client_name"])
            continue

        row_id   = generate_row_id()
        iso_date = to_iso(entry["date_obj"]) if entry["date_obj"] else None

        # Append row: 13 columns A–M
        # A=Row ID, B=Client/ID, C=Value, D-J=blank, K=Filter (client name), L=Modified/Date, M=Modified/By
        new_row = [
            row_id,                  # A: Row ID
            client_id,               # B: Client / ID
            entry["notes"],          # C: Value
            None,                    # D: Value as JSON
            None,                    # E: Edit by On
            None,                    # F: Show Original?
            None,                    # G: Original Value
            None,                    # H: Context
            None,                    # I: Temp Value (for Editing)
            None,                    # J: Change Type / Code
            entry["client_name"],    # K: Filter (client name — matches pattern of original rows)
            iso_date,                # L: Modified / Date
            entry["coach"],          # M: Modified / By
        ]
        history_ws.append(new_row)
        appended += 1

    print(f"\nSaving to {OUTPUT_FILE} ...")
    wb.save(OUTPUT_FILE)

    print("\n--- Summary ---")
    print(f"Notes processed : {len(new_notes)}")
    print(f"Notes appended  : {appended}")
    print(f"Skipped (date)  : {skipped_date}")
    print(f"Skipped (no match): {len(skipped_name)}")
    if skipped_name:
        print("\nUnmatched client names (review manually):")
        for name in sorted(set(skipped_name)):
            print(f"  - {name}")
    print("\nDone.")


if __name__ == "__main__":
    main()
