"""
import-to-cst.py
Reads All-Client-Data-Updated.xlsx and populates the CST SQLite database.

What it does:
  - Creates users: Greg Esman (admin), Shane (coach), Jeff (coach)
  - Creates 10 curriculum steps
  - Imports 339 clients with name, email, lastContactDate
  - Imports full note history from History Notes2
  - Maps coach names in notes to user IDs (Dennis/David → Greg as fallback)
  - Sets status ACTIVE by default; AT_RISK if last contact > 21 days ago

Usage:
  python3 scripts/import-to-cst.py

Requirements:
  - openpyxl (pip install openpyxl)
  - bcrypt (pip install bcrypt)
  - Run AFTER: npx prisma db push  (creates server/dev.db)
"""

import sqlite3
import uuid
import os
import re
from datetime import datetime, timezone, timedelta

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    raise

try:
    import bcrypt
except ImportError:
    print("ERROR: bcrypt not installed. Run: pip install bcrypt")
    raise

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT  = os.path.dirname(SCRIPT_DIR)
XLSX_PATH  = os.path.join(REPO_ROOT, "Client Data", "All-Client-Data-Updated.xlsx")
DB_PATH    = os.path.join(REPO_ROOT, "server", "prisma", "dev.db")

# Clients sheet (0-based column indices)
COL_CLIENT_NAME      = 0   # A
COL_CLIENT_EMAIL     = 4   # E
COL_CLIENT_NEXT_STEP = 24  # Y — most recent note text
COL_CLIENT_LAST_CONT = 32  # AG — most recent note date

# History Notes2 sheet (0-based column indices)
COL_NOTE_TEXT   = 2   # C
COL_NOTE_NAME   = 10  # K — client name
COL_NOTE_DATE   = 11  # L — ISO date
COL_NOTE_COACH  = 12  # M — coach name

CURRICULUM_STEPS = [
    "Welcome & Vision Setting",
    "Wallet Security Fundamentals",
    "Exchange Setup & KYC",
    "Liquidity Strategy 101",
    "First Deployment",
    "Advanced Yield Farming",
    "Risk Management",
    "Tax & Compliance",
    "Scaling Operations",
    "Mastermind Graduation",
]

# Active coaches — Dennis and David are no longer with us;
# their notes are kept but attributed to Greg as author.
COACH_NAME_MAP = {
    "greg esman": "greg",
    "greg":       "greg",
    "shane":      "shane",
    "jeff":       "jeff",
    "dennis":     "greg",   # former coach — attribute to Greg
    "david":      "greg",   # former coach — attribute to Greg
}

AT_RISK_DAYS = 21

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def new_id():
    return str(uuid.uuid4()).replace("-", "")[:25]

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def normalize_name(name):
    return str(name or "").strip().lower()

def parse_date(val):
    """Return a datetime (UTC) or None from an xlsx cell value."""
    if val is None:
        return None
    if isinstance(val, datetime):
        if val.tzinfo is None:
            return val.replace(tzinfo=timezone.utc)
        return val.astimezone(timezone.utc)
    s = str(val).strip()
    if not s:
        return None
    # Try ISO prefix yyyy-mm-dd
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None

def make_placeholder_email(name, idx):
    slug = re.sub(r"[^a-z0-9]", ".", normalize_name(name))
    slug = re.sub(r"\.+", ".", slug).strip(".")
    return f"{slug}.{idx}@placeholder.cst"

def hash_password(plain):
    return bcrypt.hashpw(plain.encode(), bcrypt.gensalt(10)).decode()

# ---------------------------------------------------------------------------
# Schema creation (mirrors schema.prisma for SQLite)
# ---------------------------------------------------------------------------

def create_schema(con):
    cur = con.cursor()
    cur.executescript("""
    PRAGMA journal_mode=WAL;
    PRAGMA foreign_keys=ON;

    CREATE TABLE IF NOT EXISTS users (
        id           TEXT PRIMARY KEY,
        email        TEXT UNIQUE NOT NULL,
        passwordHash TEXT NOT NULL,
        name         TEXT NOT NULL,
        role         TEXT NOT NULL DEFAULT 'COACH',
        avatar       TEXT,
        isActive     INTEGER NOT NULL DEFAULT 1,
        createdAt    TEXT NOT NULL,
        updatedAt    TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS clients (
        id                 TEXT PRIMARY KEY,
        name               TEXT NOT NULL,
        email              TEXT UNIQUE NOT NULL,
        timezone           TEXT NOT NULL DEFAULT 'UTC',
        status             TEXT NOT NULL DEFAULT 'ONBOARDING',
        riskReason         TEXT,
        onboardingStatus   TEXT NOT NULL DEFAULT 'NOT_BOOKED',
        onboardingDateTime TEXT,
        lastContactDate    TEXT,
        nextActionDate     TEXT,
        ghlContactId       TEXT,
        isDeleted          INTEGER NOT NULL DEFAULT 0,
        createdAt          TEXT NOT NULL,
        updatedAt          TEXT NOT NULL,
        coachId            TEXT NOT NULL,
        FOREIGN KEY (coachId) REFERENCES users(id)
    );

    CREATE TABLE IF NOT EXISTS client_tags (
        id        TEXT PRIMARY KEY,
        name      TEXT NOT NULL,
        clientId  TEXT NOT NULL,
        createdAt TEXT NOT NULL,
        FOREIGN KEY (clientId) REFERENCES clients(id) ON DELETE CASCADE,
        UNIQUE (clientId, name)
    );

    CREATE TABLE IF NOT EXISTS notes (
        id           TEXT PRIMARY KEY,
        content      TEXT NOT NULL,
        tags         TEXT NOT NULL DEFAULT '[]',
        isPinned     INTEGER NOT NULL DEFAULT 0,
        nextActionAt TEXT,
        createdAt    TEXT NOT NULL,
        updatedAt    TEXT NOT NULL,
        clientId     TEXT NOT NULL,
        authorId     TEXT NOT NULL,
        FOREIGN KEY (clientId) REFERENCES clients(id) ON DELETE CASCADE,
        FOREIGN KEY (authorId) REFERENCES users(id)
    );

    CREATE TABLE IF NOT EXISTS curriculum_steps (
        id        TEXT PRIMARY KEY,
        "order"   INTEGER NOT NULL,
        title     TEXT NOT NULL,
        createdAt TEXT NOT NULL,
        updatedAt TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS client_progress (
        id          TEXT PRIMARY KEY,
        isCompleted INTEGER NOT NULL DEFAULT 0,
        completedAt TEXT,
        createdAt   TEXT NOT NULL,
        updatedAt   TEXT NOT NULL,
        clientId    TEXT NOT NULL,
        stepId      TEXT NOT NULL,
        FOREIGN KEY (clientId) REFERENCES clients(id) ON DELETE CASCADE,
        FOREIGN KEY (stepId)   REFERENCES curriculum_steps(id) ON DELETE CASCADE,
        UNIQUE (clientId, stepId)
    );

    CREATE TABLE IF NOT EXISTS outcomes (
        id                 TEXT PRIMARY KEY,
        reviewStatus       TEXT NOT NULL DEFAULT 'POTENTIAL',
        reviewDone         INTEGER NOT NULL DEFAULT 0,
        endorsementStatus  TEXT NOT NULL DEFAULT 'POTENTIAL',
        endorsementCount   INTEGER NOT NULL DEFAULT 0,
        endorsementDone    INTEGER NOT NULL DEFAULT 0,
        innerCircleStatus  TEXT NOT NULL DEFAULT 'POTENTIAL',
        innerCircleDone    INTEGER NOT NULL DEFAULT 0,
        createdAt          TEXT NOT NULL,
        updatedAt          TEXT NOT NULL,
        clientId           TEXT UNIQUE NOT NULL,
        FOREIGN KEY (clientId) REFERENCES clients(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS custom_field_defs (
        id          TEXT PRIMARY KEY,
        name        TEXT NOT NULL,
        type        TEXT NOT NULL,
        options     TEXT,
        isActive    INTEGER NOT NULL DEFAULT 1,
        createdAt   TEXT NOT NULL,
        updatedAt   TEXT NOT NULL,
        createdById TEXT NOT NULL,
        FOREIGN KEY (createdById) REFERENCES users(id)
    );

    CREATE TABLE IF NOT EXISTS custom_field_values (
        id         TEXT PRIMARY KEY,
        value      TEXT NOT NULL,
        createdAt  TEXT NOT NULL,
        updatedAt  TEXT NOT NULL,
        clientId   TEXT NOT NULL,
        fieldDefId TEXT NOT NULL,
        FOREIGN KEY (clientId)   REFERENCES clients(id)           ON DELETE CASCADE,
        FOREIGN KEY (fieldDefId) REFERENCES custom_field_defs(id) ON DELETE CASCADE,
        UNIQUE (clientId, fieldDefId)
    );

    CREATE TABLE IF NOT EXISTS audit_logs (
        id           TEXT PRIMARY KEY,
        action       TEXT NOT NULL,
        resourceType TEXT,
        resourceId   TEXT,
        metadata     TEXT,
        ipAddress    TEXT,
        userAgent    TEXT,
        createdAt    TEXT NOT NULL,
        userId       TEXT,
        FOREIGN KEY (userId) REFERENCES users(id)
    );
    """)
    con.commit()
    print("  Schema ready.")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: xlsx not found at {XLSX_PATH}")
        return
    print(f"Opening {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    clients_ws = wb["Clients"]
    notes_ws   = wb["History Notes2"]

    # -----------------------------------------------------------------------
    # Connect to DB and create schema if needed
    # -----------------------------------------------------------------------
    print(f"Connecting to {DB_PATH} ...")
    con = sqlite3.connect(DB_PATH)
    create_schema(con)
    cur = con.cursor()

    # -----------------------------------------------------------------------
    # Create users (upsert by email)
    # -----------------------------------------------------------------------
    print("Creating users ...")
    users = {
        "greg": {
            "id":    new_id(),
            "email": "greg@mastermind.com",
            "name":  "Greg Esman",
            "role":  "ADMIN",
            "pw":    "changeme123",
        },
        "shane": {
            "id":    new_id(),
            "email": "shane@mastermind.com",
            "name":  "Shane",
            "role":  "COACH",
            "pw":    "changeme123",
        },
        "jeff": {
            "id":    new_id(),
            "email": "jeff@mastermind.com",
            "name":  "Jeff",
            "role":  "COACH",
            "pw":    "changeme123",
        },
    }

    for key, u in users.items():
        existing = cur.execute("SELECT id FROM users WHERE email = ?", (u["email"],)).fetchone()
        if existing:
            users[key]["id"] = existing[0]
            print(f"  User exists: {u['name']} ({existing[0]})")
        else:
            pw_hash = hash_password(u["pw"])
            ts = now_iso()
            cur.execute(
                "INSERT INTO users (id, email, passwordHash, name, role, isActive, createdAt, updatedAt) "
                "VALUES (?, ?, ?, ?, ?, 1, ?, ?)",
                (u["id"], u["email"], pw_hash, u["name"], u["role"], ts, ts)
            )
            print(f"  Created user: {u['name']} (password: {u['pw']})")

    con.commit()
    default_author_id = users["greg"]["id"]

    # -----------------------------------------------------------------------
    # Create curriculum steps (upsert by id)
    # -----------------------------------------------------------------------
    print("Creating curriculum steps ...")
    for i, title in enumerate(CURRICULUM_STEPS):
        step_id = f"step-{i+1}"
        existing = cur.execute("SELECT id FROM curriculum_steps WHERE id = ?", (step_id,)).fetchone()
        if not existing:
            ts = now_iso()
            cur.execute(
                "INSERT INTO curriculum_steps (id, title, \"order\", createdAt, updatedAt) VALUES (?, ?, ?, ?, ?)",
                (step_id, title, i, ts, ts)
            )
    con.commit()
    step_ids = [f"step-{i+1}" for i in range(len(CURRICULUM_STEPS))]
    print(f"  {len(step_ids)} curriculum steps ready")

    # -----------------------------------------------------------------------
    # Read Clients sheet → build client list
    # -----------------------------------------------------------------------
    print("Reading Clients sheet ...")
    client_rows = []
    seen_names = set()

    rows_iter = clients_ws.iter_rows(min_row=2, values_only=True)
    for row in rows_iter:
        if not row or len(row) <= COL_CLIENT_NAME:
            continue
        name = str(row[COL_CLIENT_NAME] or "").strip()
        if not name or name.lower() in seen_names:
            continue
        seen_names.add(name.lower())

        email_raw = str(row[COL_CLIENT_EMAIL] or "").strip() if len(row) > COL_CLIENT_EMAIL else ""
        last_contact_val = row[COL_CLIENT_LAST_CONT] if len(row) > COL_CLIENT_LAST_CONT else None
        last_contact = parse_date(last_contact_val)

        client_rows.append({
            "name":         name,
            "email_raw":    email_raw,
            "last_contact": last_contact,
        })

    print(f"  {len(client_rows)} unique clients found")

    # -----------------------------------------------------------------------
    # Read History Notes2 → build notes list grouped by client name
    # -----------------------------------------------------------------------
    print("Reading History Notes2 sheet ...")
    notes_by_client = {}   # normalized name → list of {text, date, coach_key}
    total_notes = 0

    for row in notes_ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= COL_NOTE_DATE:
            continue
        client_name_raw = str(row[COL_NOTE_NAME] or "").strip() if len(row) > COL_NOTE_NAME else ""
        if not client_name_raw:
            continue
        note_text = str(row[COL_NOTE_TEXT] or "").strip() if len(row) > COL_NOTE_TEXT else ""
        if not note_text:
            continue
        note_date = parse_date(row[COL_NOTE_DATE]) if len(row) > COL_NOTE_DATE else None
        coach_raw = str(row[COL_NOTE_COACH] or "").strip().lower() if len(row) > COL_NOTE_COACH else ""
        coach_key = COACH_NAME_MAP.get(coach_raw, "greg")

        key = normalize_name(client_name_raw)
        if key not in notes_by_client:
            notes_by_client[key] = []
        notes_by_client[key].append({
            "text":      note_text,
            "date":      note_date,
            "coach_key": coach_key,
        })
        total_notes += 1

    print(f"  {total_notes} notes found across {len(notes_by_client)} clients")

    # -----------------------------------------------------------------------
    # Wipe existing clients + notes (keep users + curriculum)
    # -----------------------------------------------------------------------
    print("Clearing existing client data ...")
    cur.execute("DELETE FROM audit_logs")
    cur.execute("DELETE FROM custom_field_values")
    cur.execute("DELETE FROM custom_field_defs")
    cur.execute("DELETE FROM outcomes")
    cur.execute("DELETE FROM client_progress")
    cur.execute("DELETE FROM notes")
    cur.execute("DELETE FROM client_tags")
    cur.execute("DELETE FROM clients")
    con.commit()
    print("  Cleared.")

    # -----------------------------------------------------------------------
    # Insert clients + notes
    # -----------------------------------------------------------------------
    print("Importing clients and notes ...")
    now = datetime.now(timezone.utc)
    clients_imported  = 0
    notes_imported    = 0
    used_emails       = set()

    for idx, c in enumerate(client_rows):
        name      = c["name"]
        name_key  = normalize_name(name)
        client_id = new_id()
        ts        = now_iso()

        # Email
        email = c["email_raw"]
        if not email or email.lower() in used_emails:
            email = make_placeholder_email(name, idx)
        # Ensure uniqueness
        base_email = email
        dedup = 0
        while email.lower() in used_emails:
            dedup += 1
            email = f"{base_email}.{dedup}"
        used_emails.add(email.lower())

        # Notes for this client
        client_notes = notes_by_client.get(name_key, [])

        # Derive last contact from notes if not in Clients sheet
        last_contact = c["last_contact"]
        if client_notes:
            note_dates = [n["date"] for n in client_notes if n["date"]]
            if note_dates:
                most_recent_note_date = max(note_dates)
                if last_contact is None or most_recent_note_date > last_contact:
                    last_contact = most_recent_note_date

        # Status
        if last_contact:
            days_since = (now - last_contact).days
            status = "AT_RISK" if days_since > AT_RISK_DAYS else "ACTIVE"
        else:
            status = "ACTIVE"

        # Assign coach: majority coach from notes, else Greg
        if client_notes:
            from collections import Counter
            coach_key = Counter(n["coach_key"] for n in client_notes).most_common(1)[0][0]
        else:
            coach_key = "greg"
        coach_id = users[coach_key]["id"]

        last_contact_str = last_contact.isoformat() if last_contact else None

        # Insert client
        cur.execute(
            "INSERT INTO clients "
            "(id, name, email, timezone, status, riskReason, onboardingStatus, "
            "onboardingDateTime, lastContactDate, nextActionDate, ghlContactId, "
            "isDeleted, createdAt, updatedAt, coachId) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                client_id, name, email, "UTC", status, None, "COMPLETED",
                None, last_contact_str, None, None,
                0, ts, ts, coach_id
            )
        )

        # Insert outcome record (required by schema)
        outcome_id = new_id()
        cur.execute(
            "INSERT INTO outcomes "
            "(id, reviewStatus, reviewDone, endorsementStatus, endorsementCount, "
            "endorsementDone, innerCircleStatus, innerCircleDone, createdAt, updatedAt, clientId) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (outcome_id, "POTENTIAL", 0, "POTENTIAL", 0, 0, "POTENTIAL", 0, ts, ts, client_id)
        )

        # Insert curriculum progress records
        for step_id in step_ids:
            prog_id = new_id()
            cur.execute(
                "INSERT INTO client_progress "
                "(id, isCompleted, completedAt, createdAt, updatedAt, clientId, stepId) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (prog_id, 0, None, ts, ts, client_id, step_id)
            )

        # Insert notes
        for note in client_notes:
            note_id  = new_id()
            note_ts  = note["date"].isoformat() if note["date"] else ts
            author_id = users[note["coach_key"]]["id"]
            cur.execute(
                "INSERT INTO notes "
                "(id, content, tags, isPinned, nextActionAt, createdAt, updatedAt, clientId, authorId) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (note_id, note["text"], "[]", 0, None, note_ts, note_ts, client_id, author_id)
            )
            notes_imported += 1

        clients_imported += 1
        if clients_imported % 50 == 0:
            con.commit()
            print(f"  {clients_imported}/{len(client_rows)} clients imported ...")

    con.commit()
    print()
    print("=" * 50)
    print(f"Import complete!")
    print(f"  Clients imported : {clients_imported}")
    print(f"  Notes imported   : {notes_imported}")
    print()
    print("User accounts created (all passwords: changeme123):")
    for key, u in users.items():
        print(f"  {u['role']:5}  {u['email']}  ({u['name']})")
    print()
    print("Next: start the server and log in to verify the dashboard.")

    con.close()
    wb.close()


if __name__ == "__main__":
    main()
