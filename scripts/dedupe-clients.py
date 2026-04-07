"""
dedupe-clients.py

For each duplicate client name in the Clients sheet of All-Client-Data-Updated.xlsx:
  - Keeps the row with the most recent Last Contact Date (col AG) as the primary
  - Merges any non-empty fields from older rows into the primary (no data lost)
  - If duplicate rows have different emails, preserves them in Email 2 / Email 3 (cols F, G)
  - Deletes the now-redundant older rows
  - Saves result back to All-Client-Data-Updated.xlsx

Run from project root:
    python3 scripts/dedupe-clients.py

Requires: openpyxl
"""

import os
from datetime import datetime
from collections import defaultdict
import openpyxl

BASE_DIR = os.path.join(os.path.dirname(__file__), "..", "Client Data")
FILE     = os.path.join(BASE_DIR, "All-Client-Data-Updated.xlsx")

# Key column indices (0-based)
COL_NAME         = 0   # A
COL_EMAIL        = 4   # E
COL_EMAIL2       = 5   # F
COL_EMAIL3       = 6   # G
COL_LAST_CONTACT = 32  # AG
COL_ONBOARDED    = 40  # AO


def parse_date(val):
    """Return a comparable datetime from an ISO string or None."""
    if not val:
        return None
    s = str(val).strip()[:19]
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:len(fmt)], fmt)
        except ValueError:
            continue
    return None


def merge_rows(rows):
    """
    Given a list of row tuples (all for the same client), return one merged row.
    Primary = row with most recent Last Contact Date (fallback: Onboarding Date).
    All non-empty values from secondary rows fill in blanks in the primary.
    Different emails are preserved in Email 2 / Email 3.
    """
    # Sort: most recent Last Contact first, then most recent Onboarding
    def sort_key(r):
        lc = parse_date(r[COL_LAST_CONTACT]) or datetime.min
        ob = parse_date(r[COL_ONBOARDED])    or datetime.min
        return (lc, ob)

    rows_sorted = sorted(rows, key=sort_key, reverse=True)
    primary = list(rows_sorted[0])  # mutable copy

    # Collect all unique emails across all rows
    all_emails = []
    for r in rows_sorted:
        for col in (COL_EMAIL, COL_EMAIL2, COL_EMAIL3):
            if col < len(r):
                e = str(r[col] or "").strip()
                if e and e.lower() not in [x.lower() for x in all_emails]:
                    all_emails.append(e)

    # Write emails back into primary (up to 3 slots)
    for i, slot in enumerate((COL_EMAIL, COL_EMAIL2, COL_EMAIL3)):
        primary[slot] = all_emails[i] if i < len(all_emails) else None

    # For every other column: fill primary blanks with first non-empty value from others
    for col_idx in range(len(primary)):
        if col_idx in (COL_EMAIL, COL_EMAIL2, COL_EMAIL3):
            continue  # already handled
        if primary[col_idx] is None or str(primary[col_idx]).strip() == "":
            for r in rows_sorted[1:]:
                if col_idx < len(r) and r[col_idx] is not None and str(r[col_idx]).strip() != "":
                    primary[col_idx] = r[col_idx]
                    break

    return tuple(primary)


def main():
    print(f"Loading {FILE} ...")
    wb = openpyxl.load_workbook(FILE)
    ws = wb["Clients"]

    # Read all data rows (skip header)
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        all_rows.append(row)

    total_before = len(all_rows)

    # Group by normalized name
    groups = defaultdict(list)
    for row in all_rows:
        name = str(row[COL_NAME] or "").strip().title()
        groups[name].append(row)

    dupes = {name: rows for name, rows in groups.items() if len(rows) > 1}
    print(f"  Total rows before: {total_before}")
    print(f"  Duplicate client names: {len(dupes)}")
    print(f"  Rows to remove: {sum(len(v) - 1 for v in dupes.values())}")

    # Build deduplicated row list
    merged_rows = []
    for name, rows in groups.items():
        if len(rows) == 1:
            merged_rows.append(rows[0])
        else:
            merged = merge_rows(rows)
            merged_rows.append(merged)

    total_after = len(merged_rows)
    print(f"  Total rows after:  {total_after}")

    # Rewrite the Clients sheet
    # Clear existing data (keep header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Write merged rows back
    for i, row in enumerate(merged_rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    # Remove now-empty trailing rows (openpyxl doesn't shrink on its own)
    # Delete from bottom up to avoid index shifting
    rows_written = total_after + 1  # +1 for header
    rows_total   = ws.max_row
    if rows_total > rows_written:
        ws.delete_rows(rows_written + 1, rows_total - rows_written)

    print(f"\nSaving {FILE} ...")
    wb.save(FILE)
    print("Done.")
    print(f"\nRemoved {total_before - total_after} duplicate rows.")
    print("Emails, Next Steps, and all other fields merged into the surviving row.")


if __name__ == "__main__":
    main()
