"""
process-client-data.py
Converts Client-data.xlsx into a single structured Client-data.json.

Run from the project root:
    python scripts/process-client-data.py

Requires: openpyxl
    pip install openpyxl
"""

import json
import re
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).parent.parent
XLSX_PATH = ROOT / "Client-data.xlsx"
OUT_PATH = ROOT / "Client-data.json"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize_name(raw) -> str:
    """Strip whitespace and title-case."""
    if raw is None:
        return ""
    return str(raw).strip().title()


def to_iso(value) -> str | None:
    """Convert a cell value to an ISO 8601 string, or None."""
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.isoformat()
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).isoformat()
    s = str(value).strip()
    if not s:
        return None
    # Already looks like ISO — pass through
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s
    return s  # fall back: return as-is


def cell_str(cell_value) -> str | None:
    """Return stripped string or None."""
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    return s if s else None


def get_headers(sheet) -> dict[str, int]:
    """Read the first row and return {header_lower: col_index (0-based)}."""
    headers = {}
    for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1))):
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = col_idx
    return headers


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def parse_current_data(sheet) -> dict[str, dict]:
    """
    Returns: { normalized_name -> { email, offer, current_next_step, current_milestone } }
    """
    headers = get_headers(sheet)
    print(f"  Current Data headers: {list(headers.keys())}")

    # Try to find the right columns by common names
    name_col = _find_col(headers, ["client / name", "name", "client name", "full name", "client"])
    email_col = _find_col(headers, ["client / email", "email", "email address"])
    offer_col = _find_col(headers, ["offer", "product", "program"])
    next_step_col = _find_col(headers, ["next steps / value", "next steps/value", "next steps", "next step", "value", "notes"])
    milestone_col = _find_col(headers, ["milestone name", "milestone", "current milestone"])

    print(f"  Mapped cols -> name:{name_col} email:{email_col} offer:{offer_col} "
          f"next_step:{next_step_col} milestone:{milestone_col}")

    clients = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_name = row[name_col] if name_col is not None else None
        name = normalize_name(raw_name)
        if not name:
            continue

        clients[name] = {
            "email": cell_str(row[email_col]) if email_col is not None else None,
            "offer": cell_str(row[offer_col]) if offer_col is not None else None,
            "current_next_step": cell_str(row[next_step_col]) if next_step_col is not None else None,
            "current_milestone": cell_str(row[milestone_col]) if milestone_col is not None else None,
        }

    print(f"  -> Parsed {len(clients)} clients from Current Data")
    return clients


def _sheet_has_header(sheet) -> bool:
    """
    Heuristic: if the first cell of row 1 looks like a short label (not a name/note),
    treat row 1 as a header. Otherwise, data starts at row 1.
    """
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if not first_row or first_row[0] is None:
        return True
    first_val = str(first_row[0]).strip()
    # A data row's first cell would be a person name or long text; a header is short and doesn't
    # contain common ISO date characters or newlines.
    if len(first_val) > 60 or "\n" in first_val or re.match(r"\d{4}-\d{2}", first_val):
        return False
    # If the value is a typical header keyword it's clearly a header
    for keyword in ["name", "client", "email", "note", "coach", "date", "timestamp", "step"]:
        if keyword in first_val.lower():
            return True
    # If the value looks like a proper noun (title-case, no digits, reasonable length)
    if re.match(r"^[A-Z][a-z]+(?: [A-Z][a-z]+)*$", first_val) and len(first_val) < 40:
        # Could be a client name — treat as data (no header)
        return False
    return True


def parse_call_history(sheet) -> dict[str, list]:
    """
    Returns: { normalized_name -> [ {date, notes, coach}, ... ] }
    Sorted ascending by date.
    """
    has_header = _sheet_has_header(sheet)
    start_row = 2 if has_header else 1

    if has_header:
        headers = get_headers(sheet)
        print(f"  Call History headers: {list(headers.keys())}")
        name_col = _find_col(headers, ["client / name", "name", "client name", "full name", "client"])
        notes_col = _find_col(headers, ["notes", "note", "next steps / value", "next steps/value", "next steps", "details"])
        date_col = _find_col(headers, ["timestamp", "date", "call date", "datetime"])
        coach_col = _find_col(headers, ["coach", "coach name", "assigned coach"])
    else:
        # No header — use positional columns based on what we know:
        # col 0 = name, col 1 = notes, col 2 = timestamp, col 3 = coach
        print("  Call History: no header row detected — using positional columns")
        name_col, notes_col, date_col, coach_col = 0, 1, 2, 3

    print(f"  Mapped cols -> name:{name_col} notes:{notes_col} date:{date_col} coach:{coach_col}")

    history: dict[str, list] = defaultdict(list)

    # Known stray header values that may appear as data rows
    SKIP_NAMES = {"client name", "name", "client"}

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        raw_name = row[name_col] if name_col is not None else None
        name = normalize_name(raw_name)
        if not name or name.lower() in SKIP_NAMES:
            continue

        entry = {
            "date": to_iso(row[date_col]) if date_col is not None else None,
            "notes": cell_str(row[notes_col]) if notes_col is not None else None,
            "coach": cell_str(row[coach_col]) if coach_col is not None else None,
        }
        history[name].append(entry)

    # Sort each client's history by date ascending
    for name in history:
        history[name].sort(key=lambda e: e["date"] or "")

    total = sum(len(v) for v in history.values())
    print(f"  -> Parsed {total} call entries across {len(history)} clients")
    return dict(history)


def parse_milestones(sheet) -> dict[str, list]:
    """
    Returns: { normalized_name -> [ {milestone_name, start_date, completion_date, status}, ... ] }
    """
    headers = get_headers(sheet)
    print(f"  Milestones headers: {list(headers.keys())}")

    name_col = _find_col(headers, ["client / name", "name", "client name", "full name", "client"])
    milestone_col = _find_col(headers, ["milestone name", "milestone"])
    start_col = _find_col(headers, ["start date", "start", "started"])
    completion_col = _find_col(headers, ["completion date", "completed", "completion", "end date", "end"])

    print(f"  Mapped cols -> name:{name_col} milestone:{milestone_col} "
          f"start:{start_col} completion:{completion_col}")

    milestones: dict[str, list] = defaultdict(list)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_name = row[name_col] if name_col is not None else None
        name = normalize_name(raw_name)
        if not name:
            continue

        completion_date = to_iso(row[completion_col]) if completion_col is not None else None
        status = "Completed" if completion_date else "In Progress"

        entry = {
            "milestone_name": cell_str(row[milestone_col]) if milestone_col is not None else None,
            "start_date": to_iso(row[start_col]) if start_col is not None else None,
            "completion_date": completion_date,
            "status": status,
        }
        milestones[name].append(entry)

    total = sum(len(v) for v in milestones.values())
    print(f"  -> Parsed {total} milestone entries across {len(milestones)} clients")
    return dict(milestones)


def _find_col(headers: dict[str, int], candidates: list[str]) -> int | None:
    """Return the column index for the first matching candidate header name."""
    for c in candidates:
        if c.lower() in headers:
            return headers[c.lower()]
    return None


# ---------------------------------------------------------------------------
# Name matching
# ---------------------------------------------------------------------------

def build_name_lookup(current_data: dict[str, dict]) -> dict[str, str]:
    """
    Returns { any_case_variant -> canonical_name } for all names in current_data.
    """
    lookup = {}
    for name in current_data:
        lookup[name.lower()] = name
    return lookup


def resolve_name(raw_name: str, lookup: dict[str, str]) -> str:
    """
    Try exact match first, then case-insensitive, return raw_name if no match.
    """
    if raw_name in lookup.values():
        return raw_name
    lower = raw_name.lower()
    if lower in lookup:
        return lookup[lower]
    return raw_name  # no match found — use as-is


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not XLSX_PATH.exists():
        sys.exit(f"File not found: {XLSX_PATH}")

    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print(f"Sheets found: {wb.sheetnames}")

    # -----------------------------------------------------------------------
    # Identify sheets
    # -----------------------------------------------------------------------
    sheet_names = wb.sheetnames
    current_sheet = None
    history_sheet = None
    milestones_sheet = None

    for sn in sheet_names:
        sl = sn.lower()
        if "current" in sl:
            current_sheet = wb[sn]
        elif "milestone" in sl and "history" in sl:
            milestones_sheet = wb[sn]
        elif "milestone" in sl:
            milestones_sheet = wb[sn]
        elif "history" in sl or "next step" in sl or "note" in sl:
            history_sheet = wb[sn]

    # Fallback: assign by position
    if current_sheet is None and len(sheet_names) >= 1:
        current_sheet = wb[sheet_names[0]]
    if history_sheet is None and len(sheet_names) >= 2:
        history_sheet = wb[sheet_names[1]]
    if milestones_sheet is None and len(sheet_names) >= 3:
        milestones_sheet = wb[sheet_names[2]]

    print(f"\nMapped sheets:")
    print(f"  Current Data   : {current_sheet.title if current_sheet else 'NOT FOUND'}")
    print(f"  Call History   : {history_sheet.title if history_sheet else 'NOT FOUND'}")
    print(f"  Milestones     : {milestones_sheet.title if milestones_sheet else 'NOT FOUND'}")

    # -----------------------------------------------------------------------
    # Parse
    # -----------------------------------------------------------------------
    print("\nParsing Current Data...")
    current_data = parse_current_data(current_sheet) if current_sheet else {}

    print("\nParsing Call History...")
    call_history = parse_call_history(history_sheet) if history_sheet else {}

    print("\nParsing Milestones...")
    milestones = parse_milestones(milestones_sheet) if milestones_sheet else {}

    # -----------------------------------------------------------------------
    # Build master name set
    # -----------------------------------------------------------------------
    lookup = build_name_lookup(current_data)

    all_names: set[str] = set()
    all_names.update(current_data.keys())

    # Re-key call_history using canonical names
    resolved_history: dict[str, list] = defaultdict(list)
    for name, entries in call_history.items():
        canonical = resolve_name(name, lookup)
        resolved_history[canonical].extend(entries)
    for name in resolved_history:
        resolved_history[name].sort(key=lambda e: e["date"] or "")
    all_names.update(resolved_history.keys())

    # Re-key milestones using canonical names
    resolved_milestones: dict[str, list] = defaultdict(list)
    for name, entries in milestones.items():
        canonical = resolve_name(name, lookup)
        resolved_milestones[canonical].extend(entries)
    all_names.update(resolved_milestones.keys())

    # -----------------------------------------------------------------------
    # Merge
    # -----------------------------------------------------------------------
    print(f"\nMerging {len(all_names)} unique clients...")

    client_list = []
    for name in sorted(all_names):
        cd = current_data.get(name, {})

        # Merge call history + current next step into one chronological list.
        # Historical entries are already sorted by date. The current next step
        # (from Current Data) has no date, so it is appended last as the most
        # recent note.
        history = list(resolved_history.get(name, []))
        current_note = cd.get("current_next_step")
        if current_note:
            history.append({
                "date": None,
                "notes": current_note,
                "coach": None,
            })

        client_list.append({
            "name": name,
            "email": cd.get("email"),
            "offer": cd.get("offer"),
            "current_milestone": cd.get("current_milestone"),
            "call_history": history,
            "milestones": resolved_milestones.get(name, []),
        })

    output = {"clients": client_list}

    # -----------------------------------------------------------------------
    # Write
    # -----------------------------------------------------------------------
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False, default=str)

    print(f"\nDone. Wrote {len(client_list)} clients to: {OUT_PATH}")

    # -----------------------------------------------------------------------
    # Quick stats
    # -----------------------------------------------------------------------
    in_current = sum(1 for c in client_list if c["email"] is not None)
    history_only = sum(1 for c in client_list if c["email"] is None)
    with_history = sum(1 for c in client_list if c["call_history"])
    with_current_note = sum(1 for c in client_list if c["call_history"] and c["call_history"][-1]["date"] is None)
    with_milestones = sum(1 for c in client_list if c["milestones"])

    print(f"\nStats:")
    print(f"  Total clients          : {len(client_list)}")
    print(f"  In Current Data sheet  : {in_current}")
    print(f"  History-only clients   : {history_only}")
    print(f"  Clients with call logs : {with_history}")
    print(f"  With current next step : {with_current_note}")
    print(f"  Clients with milestones: {with_milestones}")


if __name__ == "__main__":
    main()
